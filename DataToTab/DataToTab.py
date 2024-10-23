import csv
import sys
import pandas as pd
from PyQt6.QtWidgets import *
from PyQt6.QtGui import *
from PyQt6.QtCore import Qt
from Ui_DataToTab import *
from Ui_DataToTab_ReSplitDialog import *
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Font


# 类：重新拆分对话框
class ResplitDialog(QDialog, Ui_Dialog):
    def __init__(self, rawdata):
        super().__init__()
        self.setupUi(self)
        self.rawdata = rawdata

        # 按钮事件
        self.btn_ReSplit.clicked.connect(self.ReSplit)
        self.btn_Confirm.clicked.connect(self.Confirm)

    def ReSplit(self):
        newSplit = self.lineEdit.text()
        try:
            newdata = self.rawdata.replace(newSplit, "，")
            self.textEdit.setText(newdata)
        except Exception as e:
            QMessageBox.warning(self, "错误", str(e))

    def Confirm(self):
        self.rawdata = self.textEdit.toPlainText()
        self.accept()


# 类：主窗口
class MainWindows(QMainWindow, Ui_MainWindow):
    def __init__(self, parent=None):
        super(MainWindows, self).__init__(parent)
        self.setupUi(self)

        # 绑定按钮事件
        self.btn_GetCsv.clicked.connect(self.GetCsv)
        self.btn_CleanAll.clicked.connect(self.CleanAll)
        self.btn_AddLine.clicked.connect(self.AddLine)
        self.btn_DelLine.clicked.connect(self.DelLine)
        self.btn_ReSplit.clicked.connect(self.ReSplit)
        self.btn_Save.clicked.connect(self.SaveFile)

        # 初始化窗口

    # 处理文本
    def GetCsv(self):
        try:
            if self.textEdit.toPlainText() == "":
                QMessageBox.warning(self, "错误", "原始数据不能为空")
                return
            rawText = self.textEdit.toPlainText()
            lintText = rawText.split("\n")
            self.tableWidget.setRowCount(len(lintText))
            i = 0
            for text in lintText:
                item = text.split("\t")
                for j in range(4):
                    self.tableWidget.setItem(i, j, QTableWidgetItem(item[j]))
                i += 1
            self.tableWidget.resizeColumnsToContents()
            self.tabWidget.setCurrentIndex(1)
        except Exception as e:
            QMessageBox.warning(self, "错误", str(e))

    # 清空数据
    def CleanAll(self):
        self.textEdit.clear()
        self.tableWidget.setRowCount(0)
        self.tabWidget.setCurrentIndex(0)

    # 添加一行
    def AddLine(self):
        row = self.tableWidget.currentRow()
        if row == -1:
            row = self.tableWidget.rowCount()
        self.tableWidget.insertRow(row)

    # 删除一行
    def DelLine(self):
        row = self.tableWidget.currentRow()
        if row != -1:
            self.tableWidget.removeRow(row)

    # 重新拆分
    def ReSplit(self):
        rawitem = self.tableWidget.selectedItems()
        if rawitem:
            for item in rawitem:
                row = self.tableWidget.row(item)
                col = self.tableWidget.column(item)
                rawdata = item.text()
            resplitDialog = ResplitDialog(rawdata)
            if resplitDialog.exec():
                newdata = resplitDialog.rawdata
                self.tableWidget.setItem(row, col, QTableWidgetItem(newdata))
                self.tableWidget.resizeColumnsToContents()

    # 保存文件
    def SaveFile(self):
        file_path = None
        try:
            file_dialog = QFileDialog()
            file_dialog.setNameFilter("CSV文件(*.csv);;Excel文件(*.xlsx);;所有格式(*)")
            if file_dialog.exec():
                file_path = file_dialog.selectedFiles()[0]
                name_filter = file_dialog.selectedNameFilter()
                if 'Excel' in name_filter:
                    #保存为Excel
                    if not file_path.endswith('.xlsx'):
                        file_path += '.xlsx'
                else:
                    #保存为CSV
                    if not file_path.endswith(".csv"):
                        file_path += ".csv"
                try:
                    with open(file_path, "w", newline="", encoding="utf-8-sig") as csvfile:
                        csvwriter = csv.writer(csvfile)
                        for i in range(self.tableWidget.rowCount()):
                            csvwriter.writerow(
                                [
                                    self.tableWidget.item(i, 0).text(),
                                    self.tableWidget.item(i, 3).text(),
                                ]
                            )
                            csvwriter.writerow([self.tableWidget.item(i, 1).text()])
                            if self.tableWidget.item(i, 2).text() != "":
                                cn = self.tableWidget.item(i, 2).text()
                                cn_group = cn.split("，")
                                for cn in cn_group:
                                    csvwriter.writerow([cn])
                            csvwriter.writerow(
                                [
                                    "",
                                ]
                            )
                except Exception as e:
                    QMessageBox.warning(self, '错误', str(e))
                if 'Excel' in name_filter:
                    df = pd.read_csv(file_path)
                    df.to_excel(file_path, index=False)
                    #调整格式
                    wb = load_workbook(file_path)
                    ws = wb.active
                    ws.column_dimensions[get_column_letter(1)].width = 14
                    ws.column_dimensions[get_column_letter(2)].width = 8.38
                    # 描边
                    border_style = Border(
                        left=Side(border_style='thin'),
                        right=Side(border_style='thin'),
                        top=Side(border_style='thin'),
                        bottom=Side(border_style='thin')
                    )
                    # 字体
                    font_style = Font(
                        name='宋体',
                        size=11,
                        bold=False,
                        italic=False,
                        vertAlign='none',
                        underline='none',
                        strike=False,
                        color="000000"
                    )
                    row = 1
                    empty_block = 0
                    while empty_block < 2:
                        col = 1
                        if ws.cell(row=row, column=col).value is None:
                            empty_block += 1
                            if empty_block == 2:
                                wb.save(file_path)
                            row += 1
                            continue
                        else:
                            ws.cell(row=row, column=col).border = border_style
                            ws.cell(row=row, column=col).alignment = Alignment(
                                horizontal='center', vertical='center')
                            ws.cell(row=row, column=col+1).border = border_style
                            ws.cell(row=row, column=col).font = font_style
                            ws.cell(row=row, column=col+1).font = font_style
                            empty_block = 0
                            row += 1
                    wb.save(file_path)
                        
        except Exception as e:
            QMessageBox.warning(self, "错误", str(e))


if __name__ == "__main__":
    app = QApplication(sys.argv)
    mainWindows = MainWindows()
    mainWindows.show()
    sys.exit(app.exec())
